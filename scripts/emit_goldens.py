#!/usr/bin/env python3
"""Build synthetic workbook fixtures and canonical audit JSON goldens."""

from __future__ import annotations

import argparse
import filecmp
import json
import re
import shutil
import sys
import tempfile
import zipfile
from collections.abc import Callable
from datetime import UTC, datetime
from pathlib import Path

from openpyxl import Workbook

from spreadsheet_auditor.audit import audit_workbook

REPO_ROOT = Path(__file__).resolve().parents[1]
WORKBOOKS_DIR = REPO_ROOT / "tests" / "fixtures" / "workbooks"
GOLDEN_DIR = REPO_ROOT / "tests" / "fixtures" / "golden"
NORMALIZED_WORKBOOK_PATH = "<workbook>"
FIXED_TIMESTAMP = datetime(2020, 1, 1, 0, 0, 0, tzinfo=UTC)
FIXED_TIMESTAMP_TEXT = "2020-01-01T00:00:00Z"
FIXED_ZIP_TIMESTAMP = (2020, 1, 1, 0, 0, 0)
MODIFIED_RE = re.compile(
    rb"(<dcterms:modified\b[^>]*>)(.*?)(</dcterms:modified>)"
)


def canonical_report_json(workbook_path: Path) -> str:
    report = audit_workbook(workbook_path).to_dict()
    report["workbook_path"] = NORMALIZED_WORKBOOK_PATH
    return json.dumps(report, indent=2, sort_keys=True) + "\n"


def build_combined_risky(workbook: Workbook) -> None:
    sheet = workbook.active
    sheet.title = "Model"
    sheet["A1"] = 10
    sheet["A2"] = 20
    sheet["B1"] = "=A1*1.05"
    sheet["B2"] = "=TODAY()"
    sheet["B3"] = "=SUMPRODUCT(A:A,B:B)"
    sheet["B4"] = "='[source.xlsx]Sheet1'!A1"
    sheet["B5"] = "=#REF!+1"

    hidden = workbook.create_sheet("HiddenInputs")
    hidden.sheet_state = "hidden"
    hidden["A1"] = "=RAND()"


def build_formula_anomaly(workbook: Workbook) -> None:
    sheet = workbook.active
    sheet.title = "Revenue"
    sheet["A1"] = 1
    sheet["A2"] = 2
    sheet["A3"] = 3
    sheet["A4"] = 4
    sheet["A5"] = 5
    sheet["B1"] = 10
    sheet["B2"] = 20
    sheet["B3"] = 30
    sheet["B4"] = 40
    sheet["B5"] = 50
    sheet["C2"] = "=A2*B2"
    sheet["C3"] = "=A3*B3"
    sheet["C4"] = "=A4*B4+100"
    sheet["C5"] = "=A5*B5"
    sheet["C6"] = "=A6*B6"


def build_consistent_copied_formulas(workbook: Workbook) -> None:
    sheet = workbook.active
    sheet.title = "Model"
    for row in range(2, 7):
        sheet[f"A{row}"] = row - 1
        sheet[f"B{row}"] = (row - 1) * 10
        sheet[f"C{row}"] = f"=A{row}*B{row}"


def build_hidden_and_very_hidden(workbook: Workbook) -> None:
    visible = workbook.active
    visible.title = "Visible"
    visible["A1"] = "=1+1"

    hidden = workbook.create_sheet("HiddenSheet")
    hidden.sheet_state = "hidden"
    hidden["A1"] = "=TODAY()"

    very_hidden = workbook.create_sheet("Vault")
    very_hidden.sheet_state = "veryHidden"
    very_hidden["A1"] = "=RAND()"


def build_empty_workbook(workbook: Workbook) -> None:
    workbook.active.title = "Empty"


def build_single_value_cell(workbook: Workbook) -> None:
    sheet = workbook.active
    sheet.title = "Inputs"
    sheet["A1"] = "static label"


def build_static_xlsm(workbook: Workbook) -> None:
    sheet = workbook.active
    sheet.title = "MacroHost"
    sheet["A1"] = "=1+2"


def build_sheet_qualified_reference(workbook: Workbook) -> None:
    source = workbook.active
    source.title = "Source"
    source["A1"] = 42

    target = workbook.create_sheet("Target")
    target["B1"] = "=Source!A1"
    target["B2"] = "='Source'!A1*1.05"


def build_lexical_cell_ordering(workbook: Workbook) -> None:
    sheet = workbook.active
    sheet.title = "Order"
    sheet["B2"] = "=TODAY()"
    sheet["B10"] = "=TODAY()"


def build_escaping_workbook_text(workbook: Workbook) -> None:
    sheet = workbook.active
    sheet.title = "Evil<script>"
    sheet["A1"] = "=1+<script>"


FixtureBuilder = Callable[[Workbook], None]

FIXTURES: dict[str, tuple[FixtureBuilder, str]] = {
    "combined_risky": (build_combined_risky, ".xlsx"),
    "formula_anomaly": (build_formula_anomaly, ".xlsx"),
    "consistent_copied_formulas": (build_consistent_copied_formulas, ".xlsx"),
    "hidden_and_very_hidden": (build_hidden_and_very_hidden, ".xlsx"),
    "empty_workbook": (build_empty_workbook, ".xlsx"),
    "single_value_cell": (build_single_value_cell, ".xlsx"),
    "static_xlsm": (build_static_xlsm, ".xlsm"),
    "sheet_qualified_reference": (build_sheet_qualified_reference, ".xlsx"),
    "lexical_cell_ordering": (build_lexical_cell_ordering, ".xlsx"),
    "escaping_workbook_text": (build_escaping_workbook_text, ".xlsx"),
}


def write_fixture(name: str, builder: FixtureBuilder, suffix: str, output_dir: Path) -> Path:
    workbook_path = output_dir / f"{name}{suffix}"
    workbook = Workbook()
    builder(workbook)
    workbook.properties.created = FIXED_TIMESTAMP
    workbook.properties.modified = FIXED_TIMESTAMP
    workbook.save(workbook_path)
    normalize_ooxml_zip(workbook_path)
    return workbook_path


def normalize_ooxml_zip(path: Path) -> None:
    """Rewrite OOXML ZIP metadata so generated fixtures are byte-stable."""
    entries: list[tuple[zipfile.ZipInfo, bytes]] = []
    with zipfile.ZipFile(path) as workbook_zip:
        for original in workbook_zip.infolist():
            normalized = zipfile.ZipInfo(original.filename, FIXED_ZIP_TIMESTAMP)
            normalized.compress_type = zipfile.ZIP_DEFLATED
            normalized.external_attr = original.external_attr
            normalized.comment = original.comment
            content = workbook_zip.read(original.filename)
            if original.filename == "docProps/core.xml":
                content = MODIFIED_RE.sub(
                    rb"\g<1>" + FIXED_TIMESTAMP_TEXT.encode("ascii") + rb"\g<3>",
                    content,
                )
            entries.append((normalized, content))

    temp_path = path.with_suffix(f"{path.suffix}.tmp")
    with zipfile.ZipFile(temp_path, "w") as normalized_zip:
        for info, content in sorted(entries, key=lambda item: item[0].filename):
            normalized_zip.writestr(info, content)
    temp_path.replace(path)


def emit_all(output_workbooks_dir: Path, output_golden_dir: Path) -> list[tuple[Path, Path]]:
    output_workbooks_dir.mkdir(parents=True, exist_ok=True)
    output_golden_dir.mkdir(parents=True, exist_ok=True)

    emitted: list[tuple[Path, Path]] = []
    for name, (builder, suffix) in FIXTURES.items():
        workbook_path = write_fixture(name, builder, suffix, output_workbooks_dir)
        golden_path = output_golden_dir / f"{name}.json"
        golden_path.write_text(canonical_report_json(workbook_path), encoding="utf-8")
        emitted.append((workbook_path, golden_path))
    return emitted


def compare_tree(left: Path, right: Path) -> list[str]:
    errors: list[str] = []
    left_files = {path.relative_to(left) for path in left.rglob("*") if path.is_file()}
    right_files = {path.relative_to(right) for path in right.rglob("*") if path.is_file()}

    missing = sorted(left_files - right_files)
    extra = sorted(right_files - left_files)
    if missing:
        errors.append(f"missing committed files: {', '.join(str(path) for path in missing)}")
    if extra:
        errors.append(f"unexpected committed files: {', '.join(str(path) for path in extra)}")

    for relative in sorted(left_files & right_files):
        left_path = left / relative
        right_path = right / relative
        if not filecmp.cmp(left_path, right_path, shallow=False):
            errors.append(f"content differs: {relative}")
    return errors


def verify_committed() -> int:
    with tempfile.TemporaryDirectory(prefix="spreadsheet-auditor-goldens-") as temp_dir:
        temp_root = Path(temp_dir)
        generated_workbooks = temp_root / "workbooks"
        generated_goldens = temp_root / "golden"
        emit_all(generated_workbooks, generated_goldens)

        errors: list[str] = []
        errors.extend(compare_tree(generated_workbooks, WORKBOOKS_DIR))
        errors.extend(compare_tree(generated_goldens, GOLDEN_DIR))

        if errors:
            print("Golden verification failed:", file=sys.stderr)
            for error in errors:
                print(f"  - {error}", file=sys.stderr)
            print("Run `make regenerate-goldens` to refresh committed fixtures.", file=sys.stderr)
            return 1

    print("Golden verification passed.")
    return 0


def regenerate_committed() -> None:
    if WORKBOOKS_DIR.exists():
        shutil.rmtree(WORKBOOKS_DIR)
    if GOLDEN_DIR.exists():
        shutil.rmtree(GOLDEN_DIR)
    emit_all(WORKBOOKS_DIR, GOLDEN_DIR)
    print(f"Wrote {len(FIXTURES)} workbook fixtures to {WORKBOOKS_DIR}")
    print(f"Wrote {len(FIXTURES)} golden JSON files to {GOLDEN_DIR}")


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--verify",
        action="store_true",
        help="Regenerate fixtures in a temp dir and fail if committed files differ.",
    )
    args = parser.parse_args(argv)

    if args.verify:
        return verify_committed()

    regenerate_committed()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
