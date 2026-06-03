from __future__ import annotations

import re
from collections.abc import Iterable
from contextlib import suppress
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.formula import Tokenizer
from openpyxl.utils.cell import column_index_from_string

from spreadsheet_auditor.formula_pattern import (
    FormulaCellRecord,
    find_formula_pattern_anomalies,
    normalize_formula,
)
from spreadsheet_auditor.models import AuditReport, Issue, SheetSummary, build_issue

SUPPORTED_SUFFIXES = {".xlsx", ".xlsm"}
VOLATILE_FUNCTIONS = {
    "CELL",
    "INFO",
    "INDIRECT",
    "NOW",
    "OFFSET",
    "RAND",
    "RANDBETWEEN",
    "TODAY",
}
WHOLE_COLUMN_RANGE_RE = re.compile(r"(?<![A-Z0-9_])\$?[A-Z]{1,3}:\$?[A-Z]{1,3}(?![A-Z0-9_])")
EXTERNAL_WORKBOOK_RE = re.compile(r"\[[^\]]+\]")
FUNCTION_RE = re.compile(r"\b([A-Z][A-Z0-9_.]*)\s*\(")


def audit_workbook(path: str | Path) -> AuditReport:
    workbook_path = Path(path).expanduser().resolve()
    suffix = workbook_path.suffix.lower()
    if suffix not in SUPPORTED_SUFFIXES:
        supported = ", ".join(sorted(SUPPORTED_SUFFIXES))
        raise ValueError(f"Unsupported workbook format {suffix!r}; expected one of: {supported}")

    workbook = load_workbook(
        workbook_path,
        data_only=False,
        read_only=False,
        keep_vba=suffix == ".xlsm",
        keep_links=False,
    )

    sheets: list[SheetSummary] = []
    issues: list[Issue] = []

    for worksheet in workbook.worksheets:
        formula_cells: list[Cell] = []
        formula_records: list[FormulaCellRecord] = []
        for row in worksheet.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.startswith("="):
                    formula_cells.append(cell)
                    pattern = normalize_formula(value, cell.coordinate)
                    if pattern is not None:
                        formula_records.append(
                            FormulaCellRecord(
                                coordinate=cell.coordinate,
                                row=cell.row,
                                column=column_index_from_string(cell.column_letter),
                                formula=value,
                                pattern=pattern,
                            )
                        )
                    issues.extend(_lint_formula(worksheet.title, cell.coordinate, value))
                elif value == "#REF!":
                    issues.append(
                        build_issue(
                            "BROKEN_REF_VALUE",
                            message="Cell contains a broken #REF! value.",
                            sheet=worksheet.title,
                            cell=cell.coordinate,
                        )
                    )

        issues.extend(
            find_formula_pattern_anomalies(worksheet.title, formula_records)
        )

        sheets.append(
            SheetSummary(
                name=worksheet.title,
                state=worksheet.sheet_state,
                used_range=worksheet.calculate_dimension(),
                formula_cells=len(formula_cells),
            )
        )

    report = AuditReport(
        workbook_path=workbook_path,
        supported_format=suffix,
        sheets=sheets,
        issues=sorted(issues, key=_issue_sort_key),
    )
    _close_workbook(workbook)
    return report


def _close_workbook(workbook) -> None:
    workbook.close()
    vba_archive = getattr(workbook, "vba_archive", None)
    if vba_archive is not None:
        with suppress(Exception):
            vba_archive.close()


def _issue_sort_key(issue: Issue) -> tuple[str, str, str]:
    return (issue.evidence.sheet, issue.evidence.cell, issue.rule_id)


def _lint_formula(sheet: str, coordinate: str, formula: str) -> Iterable[Issue]:
    yield from _hardcoded_number_issues(sheet, coordinate, formula)
    yield from _volatile_function_issues(sheet, coordinate, formula)

    if "#REF!" in formula.upper():
        yield build_issue(
            "BROKEN_REF_FORMULA",
            message="Formula contains a broken #REF! reference.",
            sheet=sheet,
            cell=coordinate,
            formula=formula,
        )

    if WHOLE_COLUMN_RANGE_RE.search(formula.upper()):
        yield build_issue(
            "WHOLE_COLUMN_RANGE",
            message="Formula references an entire column range, which can slow recalculation.",
            sheet=sheet,
            cell=coordinate,
            formula=formula,
        )

    external_refs = sorted(set(EXTERNAL_WORKBOOK_RE.findall(formula)))
    if external_refs:
        yield build_issue(
            "EXTERNAL_WORKBOOK_REFERENCE",
            message="Formula references an external workbook.",
            sheet=sheet,
            cell=coordinate,
            formula=formula,
            details={"references": external_refs},
        )


def _hardcoded_number_issues(sheet: str, coordinate: str, formula: str) -> Iterable[Issue]:
    try:
        tokens = Tokenizer(formula).items
    except Exception as exc:  # pragma: no cover - defensive for unusual Excel grammar
        yield build_issue(
            "FORMULA_PARSE_ERROR",
            message="Formula could not be tokenized for static linting.",
            sheet=sheet,
            cell=coordinate,
            formula=formula,
            details={"error": str(exc)},
        )
        return

    constants = sorted(
        {
            token.value
            for token in tokens
            if token.type == "OPERAND" and token.subtype == "NUMBER"
        }
    )
    if constants:
        yield build_issue(
            "HARDCODED_NUMERIC_CONSTANT",
            message="Formula contains hardcoded numeric constants.",
            sheet=sheet,
            cell=coordinate,
            formula=formula,
            details={"constants": constants},
        )


def _volatile_function_issues(sheet: str, coordinate: str, formula: str) -> Iterable[Issue]:
    function_names = {
        match.group(1).split(".")[-1] for match in FUNCTION_RE.finditer(formula.upper())
    }
    volatile = sorted(function_names & VOLATILE_FUNCTIONS)
    if volatile:
        yield build_issue(
            "VOLATILE_FUNCTION",
            message="Formula uses volatile functions that can trigger expensive recalculation.",
            sheet=sheet,
            cell=coordinate,
            formula=formula,
            details={"functions": volatile},
        )
