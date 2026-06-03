from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import Workbook

from spreadsheet_auditor.audit import audit_workbook
from spreadsheet_auditor.formula_pattern import normalize_formula
from spreadsheet_auditor.models import RULES

REQUIRED_ISSUE_FIELDS = {
    "rule_id",
    "title",
    "severity",
    "category",
    "rationale",
    "remediation",
    "message",
    "evidence",
}
REQUIRED_EVIDENCE_FIELDS = {"sheet", "cell"}


def test_audit_workbook_flags_initial_static_risks(tmp_path: Path) -> None:
    workbook_path = tmp_path / "risky.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Model"
    sheet["A1"] = 10
    sheet["A2"] = 20
    sheet["B1"] = "=A1*1.05"
    sheet["B2"] = "=TODAY()"
    sheet["B3"] = "=SUMPRODUCT(A:A,B:B)"
    sheet["B4"] = "='[source.xlsx]Sheet1'!A1"
    sheet["B5"] = "=#REF!+1"

    hidden = workbook.create_sheet("HiddenInputs")
    hidden.sheet_state = "hidden"
    hidden["A1"] = "=RAND()"
    workbook.save(workbook_path)

    report = audit_workbook(workbook_path).to_dict()

    assert report["summary"] == {
        "sheet_count": 2,
        "formula_cell_count": 6,
        "issue_count": 7,
        "issues_by_severity": {"high": 1, "medium": 6},
        "issues_by_category": {
            "formula_integrity": 3,
            "lineage": 1,
            "performance": 3,
        },
    }
    assert report["sheets"][1]["state"] == "hidden"

    rule_ids = [issue["rule_id"] for issue in report["issues"]]
    assert rule_ids == [
        "VOLATILE_FUNCTION",
        "HARDCODED_NUMERIC_CONSTANT",
        "VOLATILE_FUNCTION",
        "WHOLE_COLUMN_RANGE",
        "EXTERNAL_WORKBOOK_REFERENCE",
        "BROKEN_REF_FORMULA",
        "HARDCODED_NUMERIC_CONSTANT",
    ]

    for issue in report["issues"]:
        assert REQUIRED_ISSUE_FIELDS <= set(issue)
        assert REQUIRED_EVIDENCE_FIELDS <= set(issue["evidence"])
        rule = RULES[issue["rule_id"]]
        assert issue["title"] == rule.title
        assert issue["severity"] == rule.severity
        assert issue["category"] == rule.category
        assert issue["rationale"] == rule.rationale
        assert issue["remediation"] == rule.remediation


def test_audit_report_json_is_deterministic(tmp_path: Path) -> None:
    workbook_path = tmp_path / "risky.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Model"
    sheet["B1"] = "=A1*1.05"
    sheet["B2"] = "=TODAY()"
    workbook.save(workbook_path)

    first = json.dumps(audit_workbook(workbook_path).to_dict(), indent=2, sort_keys=True)
    second = json.dumps(audit_workbook(workbook_path).to_dict(), indent=2, sort_keys=True)
    assert first == second


def test_normalize_formula_treats_relative_copies_as_same_pattern() -> None:
    first = normalize_formula("=A1+B1", "C1")
    second = normalize_formula("=A2+B2", "C2")
    assert first is not None and second is not None
    assert first == second
    absolute = normalize_formula("=$A$1+B1", "C1")
    assert absolute is not None and absolute != first


def test_anomaly_fixture_flags_single_local_outlier(tmp_path: Path) -> None:
    workbook_path = tmp_path / "anomaly.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Revenue"
    sheet["A1"] = 1
    sheet["A2"] = 2
    sheet["A3"] = 3
    sheet["A4"] = 4
    sheet["A5"] = 5
    sheet["B1"] = 10
    sheet["B2"] = 20
    sheet["B3"] = 30
    sheet["B4"] = 40
    sheet["B5"] = 50
    sheet["C2"] = "=A2*B2"
    sheet["C3"] = "=A3*B3"
    sheet["C4"] = "=A4*B4+100"
    sheet["C5"] = "=A5*B5"
    sheet["C6"] = "=A6*B6"
    workbook.save(workbook_path)

    report = audit_workbook(workbook_path)
    anomaly_issues = [
        issue for issue in report.issues if issue.rule_id == "FORMULA_PATTERN_ANOMALY"
    ]

    assert len(anomaly_issues) == 1
    issue = anomaly_issues[0]
    assert issue.evidence.sheet == "Revenue"
    assert issue.evidence.cell == "C4"
    assert issue.evidence.formula == "=A4*B4+100"
    assert issue.details["expected_pattern"] == normalize_formula("=A2*B2", "C2")
    assert issue.details["local_pattern"] == normalize_formula("=A4*B4+100", "C4")
    assert issue.details["cluster_cells"] == ["C2", "C3", "C4", "C5", "C6"]
    assert issue.details["cluster_orientation"] == "column"
    rule = RULES["FORMULA_PATTERN_ANOMALY"]
    assert issue.title == rule.title
    assert issue.remediation == rule.remediation


def test_anomaly_fixture_ignores_consistent_copied_column(tmp_path: Path) -> None:
    workbook_path = tmp_path / "consistent.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Model"
    for row in range(2, 7):
        sheet[f"A{row}"] = row - 1
        sheet[f"B{row}"] = (row - 1) * 10
        sheet[f"C{row}"] = f"=A{row}*B{row}"
    workbook.save(workbook_path)

    report = audit_workbook(workbook_path)
    anomaly_issues = [
        issue for issue in report.issues if issue.rule_id == "FORMULA_PATTERN_ANOMALY"
    ]
    assert anomaly_issues == []


def test_audit_workbook_rejects_unsupported_formats(tmp_path: Path) -> None:
    workbook_path = tmp_path / "legacy.xls"
    workbook_path.write_bytes(b"not a real workbook")

    with pytest.raises(ValueError, match="Unsupported workbook format"):
        audit_workbook(workbook_path)
