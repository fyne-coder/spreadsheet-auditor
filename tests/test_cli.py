from __future__ import annotations

import json
from html import escape
from pathlib import Path

from openpyxl import Workbook

from spreadsheet_auditor.audit import audit_workbook
from spreadsheet_auditor.cli import main
from spreadsheet_auditor.models import RULES
from spreadsheet_auditor.review_pack import render_review_pack_html


def test_scan_writes_json_report(tmp_path: Path) -> None:
    workbook_path = tmp_path / "simple.xlsx"
    output_path = tmp_path / "report.json"

    workbook = Workbook()
    workbook.active["A1"] = "=1+2"
    workbook.save(workbook_path)

    exit_code = main(["scan", str(workbook_path), "--output", str(output_path)])

    assert exit_code == 0
    report = json.loads(output_path.read_text(encoding="utf-8"))
    assert set(report) == {"workbook_path", "supported_format", "summary", "sheets", "issues"}
    assert report["summary"]["issue_count"] == 1
    issue = report["issues"][0]
    assert issue["rule_id"] == "HARDCODED_NUMERIC_CONSTANT"
    assert issue["evidence"]["cell"] == "A1"
    assert issue["evidence"]["formula"] == "=1+2"
    rule = RULES["HARDCODED_NUMERIC_CONSTANT"]
    assert issue["title"] == rule.title
    assert issue["rationale"] == rule.rationale
    assert issue["remediation"] == rule.remediation


def test_scan_writes_html_review_pack(tmp_path: Path) -> None:
    workbook_path = tmp_path / "simple.xlsx"
    review_pack_path = tmp_path / "review-pack.html"

    workbook = Workbook()
    workbook.active["A1"] = "=1+2"
    workbook.save(workbook_path)

    exit_code = main(
        ["scan", str(workbook_path), "--review-pack", str(review_pack_path)],
    )

    assert exit_code == 0
    html = review_pack_path.read_text(encoding="utf-8")
    assert "Spreadsheet Auditor Review Pack" in html
    assert escape(str(workbook_path.resolve())) in html
    assert "Workbook Summary" in html
    assert "Formula cells" in html
    assert "Issues by Severity" in html
    assert "Issues by Category" in html
    assert "HARDCODED_NUMERIC_CONSTANT" in html
    assert "A1" in html
    assert escape("=1+2") in html
    rule = RULES["HARDCODED_NUMERIC_CONSTANT"]
    assert escape(rule.remediation) in html


def test_scan_writes_json_and_html_outputs(tmp_path: Path) -> None:
    workbook_path = tmp_path / "simple.xlsx"
    json_path = tmp_path / "report.json"
    html_path = tmp_path / "review-pack.html"

    workbook = Workbook()
    workbook.active["A1"] = "=1+1"
    workbook.save(workbook_path)

    exit_code = main(
        [
            "scan",
            str(workbook_path),
            "--output",
            str(json_path),
            "--review-pack",
            str(html_path),
        ],
    )

    assert exit_code == 0
    report = json.loads(json_path.read_text(encoding="utf-8"))
    html = html_path.read_text(encoding="utf-8")
    assert report["summary"]["issue_count"] == 1
    assert "HARDCODED_NUMERIC_CONSTANT" in html


def test_review_pack_html_escapes_workbook_text(tmp_path: Path) -> None:
    workbook_path = tmp_path / "unsafe.xlsx"
    sheet_title = "Evil<script>"
    formula = "=1+<script>"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_title
    sheet["A1"] = formula
    workbook.save(workbook_path)

    report = audit_workbook(workbook_path)
    html = render_review_pack_html(report)

    assert "<script>" not in html
    assert escape(sheet_title) in html
    assert escape(formula) in html
    assert escape(str(workbook_path.resolve())) in html


def test_review_pack_export_is_deterministic(tmp_path: Path) -> None:
    workbook_path = tmp_path / "deterministic.xlsx"

    workbook = Workbook()
    workbook.active["A1"] = "=2+3"
    workbook.save(workbook_path)

    report = audit_workbook(workbook_path)
    first = render_review_pack_html(report)
    second = render_review_pack_html(report)

    assert first == second
